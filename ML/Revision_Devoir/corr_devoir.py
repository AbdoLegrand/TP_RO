import openpyxl as op
import numpy as np

wb = op.load_workbook('devoir.xlsx')
ws = wb['Feuil1']
n = 6
p = 2
Y = np.array([ws.cell(2 + i, 1).value for i in range(n)])
X = np.array([[ws.cell(2 + i, 2 + j).value for j in range(p)] for i in range(n)])

colonne_un = np.ones((n, 1))
X = np.concatenate((colonne_un, X), axis=1)
Xt = np.transpose(X)
XtX = np.dot(Xt, X)
inv = np.linalg.inv(XtX)
a = np.dot(np.dot(inv, Xt), Y)
Y_ch = np.dot(X, a)

moy = np.average(Y)
SCT = np.sum((Y - moy) ** 2)
SCE = np.sum((Y_ch - moy) ** 2)
SCR = np.sum((Y - Y_ch) ** 2)  # SCR = SCT - SCE

R2 = SCE/SCT
Ve = R2/p
VE = SCE/p
VR = SCR/(n - p - 1)
F = VE/VR


print(Ve)



