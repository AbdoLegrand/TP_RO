import numpy as np
import openpyxl

file = 'Production.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb["données"]
n = 12
p = 2
Y = np.array([ws.cell(2+i,2).value for i in range(n)])
X = np.array([ws.cell(2 + i, 3 + j).value for j in range(p) for i in range(n)])
