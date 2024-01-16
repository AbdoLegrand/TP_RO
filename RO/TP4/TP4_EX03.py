from ortools.linear_solver import pywraplp
solver = pywraplp. Solver('TP4_Exo3', pywraplp. Solver.CBC_MIXED_INTEGER_PROGRAMMING)
infinity = solver.infinity()
import openpyxl
wb=openpyxl.load_workbook('TP4_EX02.xlsx')
ws=wb['Données']
objective=solver.Objective()
X= []
n=4
m=9

A = [[ws.cell(2+i, 2+j).value for j in range(m)] for i in range (n)]
perte=[ws.cell (6,2+j).value for j in range(m)]
demande=[ws.cell(2+j, 15).value for j in range(n)]

print(perte)
print(demande)
print(A)

for i in range(m):
    x=solver.IntVar(0, infinity,'x'+str(i)) 
    X.append(x)
    
for i in range(m):
    objective.SetCoefficient(X[i], perte[i])
objective.SetMinimization()

for i in range(n):
    ctl=solver.Constraint(demande[i], infinity, 'ctl_'+str(i))
    for j in range (m):
        ctl.SetCoefficient(X[j], A[i][j])
    
solver.Solve()
# wsl=wb['rs']
# wsl.cell(1,2).value=solver.Objective().Value ()
# for j in range(m):
#     wsl.cell (3,2+j).value=int(X[j].solution_value())
# wb.save('TP4_E0X2.xlsx')
# print('Number of constraints =', solver.NumConstraints )
print('Solution:')
print('Valeur optimale =', solver.Objective().Value())

for j in range(m):
    print(int(X[j].solution_value()))
    