from ortools.linear_solver import pywraplp

solver = pywraplp.Solver('TP4_EX02', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
infinity = solver.infinity()

import openpyxl

wb = openpyxl.load_workbook('TP4_EX02.xlsx')
ws = wb.active

objective = solver.Objective()
X = []
n = 4
m = 9

A = [[ws.cell(2+i, 2+j).value for j in range(m)] for i in range(n)]
poids = [ws.cell(5, 2+j).value for j in range(m)]
demande = [ws.cell(2+j, 15).value for j in range(n)]

print(demande)

for j in range(m):
    x = solver.IntVar(0, infinity, 'x'+str(j))
    X.append(x)
    objective.SetCoefficient(X[j], 1)

print(X)

objective.SetMinimization()

for i in range(n):
    ctl = solver.Constraint(demande[i], infinity, 'ctl_'+str(i))
    for j in range(m):
        ctl.SetCoefficient(X[j], A[i][j])

solver.Solve()

print('Solution:')
print('Valeur optimale =', solver.Objective().Value())

for j in range(m):
    print(X[j].solution_value())

   