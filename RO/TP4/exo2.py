from ortools.linear_solver import pywraplp
import openpyxl

solver = pywraplp.Solver('TP3_Exo5', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
infinity = solver.infinity()

wb=openpyxl.load_workbook('Classeur1.xlsx')
ws=wb.active
objective=solver.Objective()
X=[]
n=4
m=9

A = [[ws.cell(8+i,2+j).value for j in range(m)] for i in range(n)]
Pi=[ws.cell(12,2+j).value for j in range(m)]
Di=[ws.cell(2+j,3).value for j in range(n)]

for i in range(m):
    x=solver.IntVar(0,infinity,'x'+str(i))
    X.append(x)

for i in range(m):
    objective.SetCoefficient(X[i],1)

objective.SetMinimization()

for i in range(n):
    ctl=solver.Constraint(Di[i],infinity,'ctl'+str(i))
    for j in range(m):
        ctl.SetCoefficient(X[j],A[i][j])

print('Number of constraints =', solver.NumConstraints())
solver.Solve()
print('Solution:')
print('Valeur optimale =', solver.Objective().Value())

# for j in range(m):
#     print(X[j].solution_value())