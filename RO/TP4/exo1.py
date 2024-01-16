# from ortools.linear_solver import pywraplp
# import openpyxl as op
#
#
# def solve_bin_packing_problem(file_path):
#     solver = pywraplp.Solver.CreateSolver('SCIP')
#
#     # Charger les données depuis le fichier Excel
#     wb = op.load_workbook(file_path)
#     ws = wb["Données"]
#
#     num_objects = ws.cell(3, 2).value
#     bin_capacity = ws.cell(2, 2).value
#     weights = [ws.cell(5, 2 + i).value for i in range(num_objects)]
#     print(weights)
#     print(num_objects)
#     num_bins = 20  # Fixer le nombre maximal de vols à 20 comme indiqué
#
#     # Variables
#     x = {}
#     for i in range(num_objects):
#         for j in range(num_bins):
#             x[i, j] = solver.BoolVar(f'x_{i}_{j}')
#
#     y = [solver.BoolVar(f'y_{j}') for j in range(num_bins)]
#
#     # Objective
#     solver.Minimize(solver.Sum(y))
#
#     # Constraints
#     for i in range(num_objects):
#         solver.Add(solver.Sum(x[i, j] for j in range(num_bins)) == 1)
#
#     for j in range(num_bins):
#         solver.Add(solver.Sum(weights[i] * x[i, j] for i in range(num_objects)) <= bin_capacity * y[j])
#
#     # Solve
#     status = solver.Solve()
#
#     # Write results to Excel
#     if status == pywraplp.Solver.OPTIMAL:
#         print('Optimal solution found.')
#         print('Number of bins used:', int(solver.Objective().Value()))
#
#         # Create a new workbook for results
#         # wb_result = op.Workbook()
#         ws_result = wb['Résultats']
#
#         # Write the number of bins used
#         ws_result['A1'] = 'Number of Bins Used:'
#         ws_result['B1'] = int(solver.Objective().Value())
#
#         # Write the items in each bin
#         for j in range(num_bins):
#             items_in_bin = [i for i in range(num_objects) if x[i, j].solution_value() == 1]
#             ws_result.cell(row=j + 3, column=1, value=f'Bin {j + 1}')
#             ws_result.cell(row=j + 3, column=2, value=', '.join(map(str, items_in_bin)))
#
#         # Save the results to a new Excel file
#         wb.save('Interface_TP3_Exo1.xlsx')
#
#     else:
#         print('The problem does not have an optimal solution.')
#
#
# # Example usage
# solve_bin_packing_problem('Interface_TP3_Exo1.xlsx')  # Remplacez 'input_data.xlsx' par le chemin de votre fichier Excel

from ortools.linear_solver import pywraplp
import openpyxl as op

file_path = 'Exo1_tp4.xlsx'
wb = op.load_workbook(file_path)
ws = wb['Feuil1']

weights = [ws.cell(2, i+1).value for i in range(1, 11)]
quantities = [ws.cell(3, i+1).value for i in range(1, 11)]
bin_capacity = 100
max_bins = 20
solver = pywraplp.Solver.CreateSolver('SCIP')

x = {}
for i in range(len(weights)):
    for j in range(max_bins):
        x[i, j] = solver.IntVar(0, solver.infinity(), 'x[%i,%i]' % (i, j))
y = {}
for j in range(max_bins):
    y[j] = solver.IntVar(0, 1, 'y[%i]' % j)

# Chaque objet doit être complètement assigné à des bins
for i in range(len(weights)):
    solver.Add(sum(x[i, j] for j in range(max_bins)) == int(quantities[i]))

# La capacité des bins ne doit pas être dépassée
for j in range(max_bins):
    solver.Add(sum(x[i, j] * weights[i] for i in range(len(weights))) <= y[j] * bin_capacity)

objective = solver.Objective()
for j in range(max_bins):
    objective.SetCoefficient(y[j], 1)
objective.SetMinimization()

# Résolution
status = solver.Solve()

# Écrire les résultats dans un fichier Excel
result_wb = op.Workbook()
result_ws = result_wb.active

# Entête
result_ws.cell(row=1, column=1, value="Bin")
result_ws.cell(row=1, column=2, value="Objet")
result_ws.cell(row=1, column=3, value="Quantité")

# Remplissage des résultats
if status == pywraplp.Solver.OPTIMAL:
    print('Solution trouvée !')
    row_num = 2
    for j in range(max_bins):
        if y[j].solution_value() > 0:
            for i in range(len(weights)):
                if x[i, j].solution_value() > 0:
                    result_ws.cell(row=row_num, column=1, value=j + 1)
                    result_ws.cell(row=row_num, column=2, value=i + 1)
                    result_ws.cell(row=row_num, column=3, value=x[i, j].solution_value())
                    row_num += 1
    result_ws.cell(row=row_num, column=1, value="Solution optimal")
    result_ws.cell(row=row_num, column=2, value=sum(y[j].solution_value() for j in range(max_bins)))
    row_num += 1
# Enregistrement du fichier Excel résultant
result_wb.save('resultats4.xlsx')
