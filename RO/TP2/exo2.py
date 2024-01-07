#from __future__ import print_function
from ortools.linear_solver import pywraplp
import openpyxl as op

def TP2_EX02():
    
    solver = pywraplp.Solver('TP2_EX02',pywraplp.Solver.GLOP_LINEAR_PROGRAMMING)
    
    file = 'Interface_TP2_Exo2.xlsx'
    wb=op.load_workbook(file)
    ws=wb["input"]
    
    # Création des variables réelles x1,...x7   
    X=[]
    ct_stock=[]
    for i in range(2,9):
        ct_stock.append(ws.cell(i,5).value)
    print(ct_stock)
    for i in range(7):    
        x=solver.NumVar(0, ct_stock[i], 'x_'+str(i))
        X.append(x)
        
    # Contrainte de satisfaction de la demande
    demandeMax=5000
    demandeMin=5000
    ct_demande=solver.Constraint(demandeMin,demandeMax,'ct_demande')
    for i in range(7):
        ct_demande.SetCoefficient(X[i],1)

    # Contrainte de respect du pourcentage de carbone
    min_carb=(ws.cell(3,9).value)*demandeMax
    max_carb=(ws.cell(3,10).value)*demandeMax
    ct_carb=solver.Constraint(min_carb,max_carb,'ct_carb')
    Carb=[]
    for i in range(2,9):
       Carb.append(ws.cell(i,2).value)
    
    print(Carb)
    for i in range(7):    
       ct_carb.SetCoefficient(X[i] ,Carb[i])
        
    # Contrainte de respect du pourcentage de cuivre
       
    min_cuiv=(ws.cell(4,9).value)*demandeMax
    max_cuiv=(ws.cell(4,10).value)
    max_cuiv = float(max_cuiv)*demandeMax
    

    ct_cuiv=solver.Constraint(min_cuiv,max_cuiv,'ct_cuiv')
    Cuiv=[]
    for i in range(2,9):
        Cuiv.append(ws.cell(i,3).value)
    
    print(Cuiv)
    for i in range(7):
        ct_cuiv.SetCoefficient(X[i], Cuiv[i])
   

    # Contrainte de respect du pourcentage de manganèze
    min_mang=(ws.cell(5,9).value)*demandeMax
    max_mang=(ws.cell(5,10).value)*demandeMax
    ct_mang=solver.Constraint(min_mang,max_mang,'ct_mang')
    Mang=[]
    for i in range(2,9):
       Mang.append(ws.cell(i,4).value)
    
    print(Mang)
    for i in range(7):
        ct_mang.SetCoefficient(X[i], Mang[i])

    # Minimizer 1.2*x1 + 1.5*x2 + 0.9*x3 + 1.3*x4 + 1.45*x5 + 1.2*x6 + 1*x7.
    objective= solver.Objective()
    coef=[]
    for i in range(2,9):
       coef.append(ws.cell(i,6).value)

    print(coef)
    for i in range(7):
        objective.SetCoefficient(X[i], coef[i])
        objective.SetMinimization()

    solver.Solve()

    print('\nSolution:')
    print('Objective value =', solver.Objective().Value())

    for i in range(7):
        print('x_'+str(i)+'=', X[i].solution_value())
    
TP2_EX02()