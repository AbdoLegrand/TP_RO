# import openpyxl
# from ortools.linear_solver import pywraplp
#
# # Chargement du classeur Excel
# wb = openpyxl.load_workbook('file_exo1.xlsx')
#
# # Lecture des données à partir de la feuille "Données"
# ws = wb['Données']
# n = ws.cell(1, 21).value  # Nombre d'objets
# capacity = ws.cell(2, 2).value  # Capacité maximale
# print(ws.cell(3, 3).value)
# benefice = [ws.cell(2, 2 + i).value for i in range(n)]  # Liste des bénéfices
# poids = [ws.cell(3, 2 + i).value for i in range(n)]  # Liste des poids
#
# # Création du solveur
# solver = pywraplp.Solver.CreateSolver('SCIP')
#
# # Variables de décision
# x = [solver.BoolVar('x[%i]' % i) for i in range(n)]
#
# # Contrainte de capacité
# constraint = solver.Constraint(0, capacity)
# for i in range(n):
#     constraint.SetCoefficient(x[i], poids[i])
#
# # Fonction objectif
# objective = solver.Objective()
# for i in range(n):
#     objective.SetCoefficient(x[i], benefice[i])
# objective.SetMaximization()
#
# # Résolution du problème
# solver.Solve()
#
# # Affichage des résultats
# ws = wb['Résultats']
# ws['B1'] = objective.Value()
#
# for i in range(n):
#     ws.cell(3, 2 + i).value = x[i].solution_value()
#
# # Sauvegarde du classeur Excel
# wb.save('file_exo1.xlsx')


from ortools.linear_solver import pywraplp
import openpyxl as op
def TP3_Exo1():
    solver = pywraplp.Solver('TP3_Exo1', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    # infinity = solver.infinity()
    file = 'Interface_TP3_Exo1.xlsx'
    wb = op.load_workbook(file)
    ws = wb["Données"]
    n = ws.cell(3,2).value
    capacity = ws.cell(2,2).value
    benefice = [ws.cell(4,2+i).value for i in range(n)]
    poids = [ws.cell(5,2+i).value for i in range(n)]

    X=[]
    for i in range(n):
        x= solver.IntVar(0,1, 'x_'+str(i))
        X.append(x)
    objective =solver.Objective()

    for i in range(n):
        objective.SetCoefficient(X[i],benefice[i])
        objective.SetMaximization()
    ct = solver.Constraint(0, capacity, 'ct')
    for i in range(n):
        ct.SetCoefficient(X[i], poids[i])
    solver.Solve()
    print('Solution:')
    print('Valeur optimale =', solver.Objective().Value())
    ws = wb["Résultats"]
    ws['B1'] = "objValue"
    for i in range(n):
        ws.cell(3,2+i).value = X[i].solution_value()
    wb.save(file)
TP3_Exo1()

