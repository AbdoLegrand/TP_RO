from ortools.linear_solver import pywraplp
import openpyxl as op


def TP3_Exo3():
    solver = pywraplp.Solver('TP3_Exo3', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()
    file = 'Interface_TP3_Exo.xlsx'
    wb = op.load_workbook(file)
    ws = wb["Données"]
    m = 10
    n = 20
    capacity = 100000
    benefice = [ws.cell(2, 2 + i).value for i in range(n)]
    poids = [ws.cell(3, 2 + i).value for i in range(n)]
    Nombre = [ws.cell(4, 2 + i).value for i in range(n)]

    # Définition des variables de décision
    X = []
    for i in range(n):
        y = []
        for j in range(m):
            x = solver.IntVar(0, Nombre[i], 'x_' + str(i) + '_' + str(j))
            y.append(x)
        X.append(y)
    objective = solver.Objective()

    # Définition de la fonction objectif
    for i in range(n):
        for j in range(m):
            objective.SetCoefficient(X[i][j], benefice[i])
    objective.SetMaximization()

    for j in range(m):
        ct = solver.Constraint(0, capacity, 'ct' + str(j))
        for i in range(n):
            ct.SetCoefficient(X[i][j], poids[i])

    for i in range(n):
        ctt = solver.Constraint(0, Nombre[i], 'ctt' + str(i))
        for j in range(m):
            ctt.SetCoefficient(X[i][j], 1)

    solver.Solve()
    print('Solution:')
    print('Valeur optimale =', solver.Objective().Value())


TP3_Exo3()