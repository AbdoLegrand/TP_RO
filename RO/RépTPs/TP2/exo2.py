#from __future__ import print_function
from ortools.linear_solver import pywraplp
import openpyxl as op

def exo2():
    
    solver = pywraplp.Solver('exo2' ,pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    
    file = 'Interface_TP2_Exo2.xlsx'
    wb = op.load_workbook(file)
    ws = wb["input"]
    n = 7
    D = ws.cell(12, 4).value
    obj = [ws.cell(i + 2, 6).value for i in range(n)]

    min_cr = ws.cell(3, 9).value
    max_cr = ws.cell(3, 10).value
    cr = [ws.cell(2 + i, 2).value for i in range(n)]


    min_cu = ws.cell(4, 9).value
    max_cu = ws.cell(4, 10).value
    cu = [ws.cell(2 + i, 3).value for i in range(n)]

    min_mang = ws.cell(5, 9).value
    max_mang = ws.cell(5, 10).value
    mang = [ws.cell(2 + i, 4).value for i in range(n)]

    X = []
    max_x = [ws.cell(2 + i, 5).value for i in range(n)]

    for i in range(n):
        x = solver.NumVar(0, max_x[i], 'x'+str(i+1))
        X.append(x)
    objective = solver.Objective()

    for i in range(n):
        objective.SetCoefficient(X[i], obj[i])
    objective.SetMinimization()

    ct_D = solver.Constraint(D, D, 'ct_D')
    ct_cr = solver.Constraint(min_cr*D, max_cr*D, 'ct_cr')
    ct_cu = solver.Constraint(min_cu*D, max_cu*D, 'ct_cu')
    ct_mang = solver.Constraint(min_mang*D, max_mang*D, 'ct_mang')

    for i in range(n):
        ct_D.SetCoefficient(X[i], 1)
        ct_cr.SetCoefficient(X[i], cr[i])
        ct_cu.SetCoefficient(X[i], cu[i])
        ct_mang.SetCoefficient(X[i], mang[i])

    solver.Solve()
    print('Number of variables =', solver.NumVariables())
    print('Number of constraints =', solver.NumConstraints())
    print('Solution:')
    print('Valeur optimale =', solver.Objective().Value())

    for i in range(7):
        print('x', i, ' = ', X[i].solution_value())

    print('Advanced usage:')
    print('Problem solved in %f milliseconds' % solver.wall_time())
    print('Problem solved in %d iterations' % solver.iterations())
    print('Problem solved in %d branch-and-bound nodes' % solver.nodes())

exo2()