from ortools.linear_solver import pywraplp
import openpyxl 
wb=openpyxl.load_workbook('EXO3.xlsx')
ws=wb.active

def EX03():
    solver = pywraplp.Solver('Exo3',  pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()

    X = []
    aliment = 2
    mp = 3
    D = [9000,12000]
    protein = []
    for i in range(2, 5):
        protein.append(ws.cell(2, i).value)
    print(protein)    
 
    lipide=[]
    for i in range(2,5):
        lipide.append(ws.cell(3,i).value)
    glucide=[]
    for i in range(2,5):
        glucide.append(ws.cell(4,i).value)
    
    stocks=[]
    for i in range(2,5):
        stocks.append(ws.cell(9,i).value)
    
    prix=[]
    for i in range(2,5):
         prix.append(ws.cell(10,i).value)
    
    for i in range (aliment):
        U = []
        for j in range(mp):
            x = solver.NumVar(0.0, infinity, 'x_' + str(i) + '_' + str(j))
            U.append(x)
        X.append(U)
    print('Number of variable =', solver.NumVariables())

    print(X)
    # le probleme est l'equation objective
    objective = solver.Objective()
    for i in range (mp):
        for j in range(aliment):
            objective.SetCoefficient(X[j][i],prix[i])
            # print(X[j][i],' * ',PK[i])
    for i in range(mp-1):
        for n in range(aliment):
            objective.SetCoefficient(X[i][n],1.5)
            # print(X[i][n],' * ',1.5)

    # print(objective.)
    objective.SetMinimization()

    for i in range(aliment):
        ct_D = solver.Constraint(D[i], D[i], 'ct_D_' + str(i))
        for j in range(mp):
            ct_D.SetCoefficient(X[i][j], 1)

    for i in range(aliment):
        ct_P = solver.Constraint((9.5*D[i]), infinity, 'ct_P_' + str(i))
        for j in range(mp):
            ct_P.SetCoefficient(X[i][j], protein[j])
            print(X[i][j],' * ',protein[j])


    for i in range(aliment):
        ct_L = solver.Constraint((2 * D[i]), infinity, 'ct_L_' + str(i))
        for j in range(mp):
            ct_L.SetCoefficient(X[i][j], lipide[j])


    for i in range(aliment):
        ct_G = solver.Constraint(0, (6 * D[i]), 'ct_G_' + str(i))
        for j in range(mp):
            ct_G.SetCoefficient(X[i][j], glucide[j])


    for i in range(0, len(stocks)):
        ct_S = solver.Constraint(0, stocks[i], 'ct_S_' + str(i))
        for j in range(aliment):
            ct_S.SetCoefficient(X[j][i], 1)




    print('Nombre des contraintes =', solver.NumConstraints())
    solver.Solve()
    print('Solution:')
    print('Valeur optimale =', solver.Objective().Value())

EX03()