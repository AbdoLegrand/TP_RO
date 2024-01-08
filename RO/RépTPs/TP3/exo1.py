from ortools.linear_solver import pywraplp
import openpyxl as op
def TP3_Exo1():

    solver = pywraplp.Solver('TP3_Exo1', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
    infinity = solver.infinity()
    file = 'Interface_TP3_Exo1.xlsx'
    wb = op.load_workbook(file)
    ws = wb["Données"]
    n = ws.cell(1,2).value
    capacity = ws.cell(2,2).value
    benefice = [ws.cell(4,2+i).value for i in range(n)]
    poids = [ws.cell(5,2+i).value for i in range(n)]

    X=[]
    for i in range(n):
        x= solver.IntVar(0,1, 'x_'+str(i))
        X.append(x)
    objective =solver.Objective()

    for i in range(n):
        objective.SetCoefficient(X[i],benefice[i])
    objective.SetMaximization()
    ct = solver.Constraint(0, capacity, 'ct')
    for i in range(n):
        ct.SetCoefficient(X[i], poids[i])
    solver.Solve()
    print("Number of variables : ", solver.NumVariables())
    print("Number of constraints : ", solver.NumConstraints())
    print('Solution:')
    print('Valeur optimale =', solver.Objective().Value())
    ws = wb["Résultats"]
    ws['B1'] = "objValue"
    for i in range(n):
        ws.cell(3, 2+i).value = X[i].solution_value()
    wb.save(file)


TP3_Exo1()
