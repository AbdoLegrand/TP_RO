# from ortools.linear_solver import pywraplp
# import openpyxl as op
#
#
# def TP3_Exo3():
#     solver = pywraplp.Solver('TP3_Exo3', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)
#     infinity = solver.infinity()
#     file = 'Interface_TP3_Exo3.xlsx'
#     wb = op.load_workbook(file)
#     ws = wb["Données"]
#
#     m = 10
#     n = ws.cell(1, 2).value
#     capacity = ws.cell(2, 2).value
#     benefice = [ws.cell(4, 2 + i).value for i in range(n)]
#     poids = [ws.cell(5, 2 + i).value for i in range(n)]
#     Nombre = [ws.cell(6, 2 + i).value for i in range(n)]
#     X = []
#     for i in range(n):
#         y = []
#         for j in range(m):
#             x = solver.IntVar(0, Nombre[i], 'x_' + str(i) + '_' + str(j))
#             y.append(x)
#         X.append(y)
#     objective = solver.Objective()
#
#     for i in range(n):
#         for j in range(m):
#             objective.SetCoefficient(X[i][j], benefice[i])
#     objective.SetMaximization()
#
#     for j in range(m):
#         ct = solver.Constraint(0, capacity, 'ct' + str(j))
#         for i in range(n):
#             ct.SetCoefficient(X[i][j], poids[i])
#     for i in range(n):
#         ctt = solver.Constraint(0, Nombre[i], 'ctt' + str(i))
#         for j in range(m):
#             ctt.SetCoefficient(X[i][j], 1)
#
#     solver.Solve()
#     print('Solution:')
#     print('Valeur optimale =', solver.Objective().Value())
#
#
# TP3_Exo3()


from ortools.linear_solver import pywraplp
import openpyxl as op


def TP2_EX02():
    solver = pywraplp.Solver('TP2_EX02', pywraplp.Solver.CBC_MIXED_INTEGER_PROGRAMMING)

    file = 'Interface_TP3_Exo3.xlsx'
    wb = op.load_workbook(file)

    ws = wb["Données"]
    n = ws.cell(1, 2).value
    capacity = ws.cell(2, 2).value
    benefice = [ws.cell(4, 2 + i).value for i in range(n)]
    poids = [ws.cell(5, 2 + i).value for i in range(n)]
    nombre = [ws.cell(6, 2 + i).value for i in range(n)]

    X = []

    objective = solver.Objective()

    for i in range(n):
        U = []
        ct_capacite = solver.Constraint(0, capacity, 'ct_capacite_' + str(i))
        for j in range(10):
            x = solver.IntVar(0.0, nombre[i], 'x_' + str(i) + '_' + str(j))
            U.append(x)
            ct_capacite.SetCoefficient(x, poids[i])
            objective.SetCoefficient(x, benefice[i])
        X.append(U)

    objective.SetMaximization()

    solver.Solve()

    print('Solution:')
    print('Valeur optimale =', solver.Objective().Value())
    # for i in range(n):
    #     for j in range(10):
    #         print(f"x_{i}_{j}: {X[i][j].solution_value()}")

    # Uncomment the following lines if you want to save the results back to the Excel file
    ws = wb["Résultats"]
    ws['B1'] = "Solution Optimale :"
    ws['C1'] = solver.Objective().Value()
    for i in range(n):
        ws.cell(2, 2+i).value = X[i][j].solution_value()
    wb.save(file)


TP2_EX02()
